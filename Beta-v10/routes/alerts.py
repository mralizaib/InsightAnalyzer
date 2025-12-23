from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for
from flask_login import login_required, current_user
import logging
from datetime import datetime, timedelta
import json

logger = logging.getLogger(__name__)

alerts_bp = Blueprint('alerts', __name__)

@alerts_bp.route('/alerts')
@login_required
def index():
    # Get user's alert configurations with error handling
    alerts = []
    try:
        from models import AlertConfig
        alerts = AlertConfig.query.filter_by(user_id=current_user.id).all()
    except Exception as e:
        logger.error(f"Error loading alert configurations: {str(e)}")
        flash('Unable to load alert configurations. Database may need migration.', 'warning')
    return render_template('alerts.html', alerts=alerts)

@alerts_bp.route('/api/alerts', methods=['GET'])
@login_required
def get_alerts():
    """Get security alerts from OpenSearch"""
    try:
        from opensearch_api import OpenSearchAPI
        opensearch = OpenSearchAPI()

        # Get query parameters
        severity_levels = request.args.getlist('severity_levels[]')
        if not severity_levels:
            severity_levels = ['critical', 'high']  # Default levels

        # Time range
        time_range = request.args.get('time_range', '24h')
        end_time = datetime.utcnow().isoformat()

        if time_range == '1h':
            start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()
        elif time_range == '6h':
            start_time = (datetime.utcnow() - timedelta(hours=6)).isoformat()
        elif time_range == '24h':
            start_time = (datetime.utcnow() - timedelta(days=1)).isoformat()
        elif time_range == '7d':
            start_time = (datetime.utcnow() - timedelta(days=7)).isoformat()
        elif time_range == '30d':
            start_time = (datetime.utcnow() - timedelta(days=30)).isoformat()
        elif time_range == '60d':
            start_time = (datetime.utcnow() - timedelta(days=60)).isoformat()
        elif time_range == '90d':
            start_time = (datetime.utcnow() - timedelta(days=90)).isoformat()
        else:
            # Custom time range
            start_time = request.args.get('start_time')
            end_time = request.args.get('end_time', end_time)

        # Pagination
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))

        # Other filters
        sort_field = request.args.get('sort_field', '@timestamp')
        sort_order = request.args.get('sort_order', 'desc')

        # Additional filters
        additional_filters = {}
        search_query = request.args.get('search_query')
        if search_query:
            additional_filters['search_query'] = search_query  # Use search_query key for multi-field search

        rule_id = request.args.get('rule_id')
        if rule_id:
            additional_filters['rule.id'] = rule_id

        # FIM alerts filter
        fim_alerts = request.args.get('fim_alerts')
        if fim_alerts == 'true':
            additional_filters['rule.id'] = ['553', '554']

        # Search for alerts
        results = opensearch.search_alerts(
            severity_levels=severity_levels,
            start_time=start_time,
            end_time=end_time,
            limit=limit,
            offset=offset,
            sort_field=sort_field,
            sort_order=sort_order,
            additional_filters=additional_filters
        )

        return jsonify(results)
    except Exception as e:
        logger.error(f"Error getting alerts: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alert_configs', methods=['GET'])
@login_required
def get_alert_configs():
    """Get all alert configurations for the current user"""
    try:
        from models import AlertConfig
        alerts = AlertConfig.query.filter_by(user_id=current_user.id).all()

        alerts_list = []
        for alert in alerts:
            try:
                alert_data = {
                    'id': alert.id,
                    'name': alert.name,
                    'alert_levels': alert.get_alert_levels(),
                    'email_recipient': alert.email_recipient,
                    'notify_time': alert.notify_time,
                    'enabled': alert.enabled,
                    'created_at': alert.created_at.isoformat()
                }

                # Add include_fields if available
                try:
                    if hasattr(alert, 'get_include_fields') and callable(getattr(alert, 'get_include_fields')):
                        alert_data['include_fields'] = alert.get_include_fields()
                    else:
                        alert_data['include_fields'] = ["@timestamp", "agent.ip", "agent.labels.location.set", "agent.name", "rule.description", "rule.id"]
                except Exception as field_error:
                    logger.warning(f"Error getting include_fields for alert {alert.id}: {str(field_error)}")
                    alert_data['include_fields'] = ["@timestamp", "agent.ip", "agent.labels.location.set", "agent.name", "rule.description", "rule.id"]

                alerts_list.append(alert_data)
            except Exception as alert_error:
                logger.error(f"Error processing alert config {alert.id}: {str(alert_error)}")
                continue

        return jsonify(alerts_list)
    except Exception as e:
        logger.error(f"Error getting alert configs: {str(e)}")
        # Return empty list instead of error to allow page to load
        return jsonify([])

@alerts_bp.route('/api/alert_configs', methods=['POST'])
@login_required
def create_alert_config():
    """Create a new alert configuration"""
    from models import AlertConfig
    from flask import current_app
    
    db = None
    try:
        db = current_app.extensions['sqlalchemy'].db
        
        data = request.json

        # Validate required fields
        if not data.get('name'):
            return jsonify({'error': 'Alert name is required'}), 400

        if not data.get('alert_levels'):
            return jsonify({'error': 'At least one alert level must be selected'}), 400

        if not data.get('email_recipient'):
            return jsonify({'error': 'Email recipient is required'}), 400

        # Create new alert configuration
        new_alert = AlertConfig(
            user_id=current_user.id,
            name=data.get('name'),
            email_recipient=data.get('email_recipient'),
            notify_time=data.get('notify_time'),
            enabled=data.get('enabled', True)
        )

        # Set JSON fields
        new_alert.set_alert_levels(data.get('alert_levels'))

        # Set include_fields if provided
        if 'include_fields' in data and hasattr(new_alert, 'set_include_fields'):
            new_alert.set_include_fields(data.get('include_fields'))

        # Save to database
        db.session.add(new_alert)
        db.session.commit()

        return jsonify({
            'id': new_alert.id,
            'name': new_alert.name,
            'message': 'Alert configuration created successfully'
        }), 201
    except Exception as e:
        if db:
            try:
                db.session.rollback()
            except:
                pass
        logger.error(f"Error creating alert config: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>', methods=['PUT'])
@login_required
def update_alert_config(alert_id):
    """Update an existing alert configuration"""
    from models import AlertConfig
    from flask import current_app
    
    db = None
    try:
        db = current_app.extensions['sqlalchemy'].db
        
        # Validate request data
        if not request.is_json:
            return jsonify({'error': 'Request must be JSON'}), 400
            
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        # Validate required fields
        if 'name' in data and not data['name'].strip():
            return jsonify({'error': 'Alert name cannot be empty'}), 400
            
        if 'alert_levels' in data and not data['alert_levels']:
            return jsonify({'error': 'At least one alert level must be selected'}), 400
            
        if 'email_recipient' in data and not data['email_recipient'].strip():
            return jsonify({'error': 'Email recipient cannot be empty'}), 400

        # Update fields if provided
        if 'name' in data:
            alert.name = data['name'].strip()

        if 'alert_levels' in data:
            alert.set_alert_levels(data['alert_levels'])

        if 'email_recipient' in data:
            alert.email_recipient = data['email_recipient'].strip()

        if 'notify_time' in data:
            alert.notify_time = data['notify_time']

        if 'enabled' in data:
            alert.enabled = data['enabled']

        # Set include_fields if provided
        if 'include_fields' in data and hasattr(alert, 'set_include_fields'):
            alert.set_include_fields(data.get('include_fields'))

        # Save changes
        db.session.commit()

        return jsonify({
            'id': alert.id,
            'message': 'Alert configuration updated successfully'
        })
    except Exception as e:
        if db:
            try:
                db.session.rollback()
            except:
                pass
        logger.error(f"Error updating alert config: {str(e)}")
        return jsonify({'error': f'Failed to update alert configuration: {str(e)}'}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_alert_config(alert_id):
    """Delete an alert configuration"""
    from models import AlertConfig
    from flask import current_app
    
    db = None
    try:
        db = current_app.extensions['sqlalchemy'].db
        
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        # Store alert name for response
        alert_name = alert.name
        
        db.session.delete(alert)
        db.session.commit()

        return jsonify({
            'message': f'Alert configuration "{alert_name}" deleted successfully'
        })
    except Exception as e:
        if db:
            try:
                db.session.rollback()
            except:
                pass
        logger.error(f"Error deleting alert config: {str(e)}")
        return jsonify({'error': f'Failed to delete alert configuration: {str(e)}'}), 500

@alerts_bp.route('/api/alerts/<string:alert_id>', methods=['GET'])
@login_required
def get_alert_details(alert_id):
    """Get details for a specific alert by ID"""
    try:
        from opensearch_api import OpenSearchAPI
        
        index = request.args.get('index')
        opensearch = OpenSearchAPI()

        result = opensearch.get_alert_by_id(alert_id, index)

        if 'error' in result:
            logger.error(f"Error getting alert details: {result['error']}")
            return jsonify(result), 404

        return jsonify(result)
    except Exception as e:
        logger.error(f"Error loading alert details: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>/debug', methods=['GET'])
@login_required
def debug_alert(alert_id):
    """Debug an alert configuration to see why it might not be triggering"""
    try:
        from models import AlertConfig
        from opensearch_api import OpenSearchAPI
        from email_alerts import EmailAlerts
        
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        debug_info = {
            'alert_config': {
                'id': alert.id,
                'name': alert.name,
                'enabled': alert.enabled,
                'notify_time': alert.notify_time,
                'email_recipient': alert.email_recipient,
                'alert_levels': alert.get_alert_levels()
            },
            'current_time': {
                'utc': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                'pakistan': (datetime.utcnow() + timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')
            },
            'recent_alerts': [],
            'sent_alerts': [],
            'smtp_config': {
                'server': 'configured' if hasattr(EmailAlerts(), 'smtp_server') else 'not configured',
                'username': 'configured' if hasattr(EmailAlerts(), 'smtp_username') and EmailAlerts().smtp_username else 'not configured'
            }
        }

        # Check for recent alerts
        opensearch = OpenSearchAPI()
        end_time = datetime.utcnow().isoformat()
        start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()

        alerts_data = opensearch.search_alerts(
            severity_levels=alert.get_alert_levels(),
            start_time=start_time,
            end_time=end_time,
            limit=10
        )

        if 'results' in alerts_data:
            debug_info['recent_alerts'] = [
                {
                    'timestamp': a.get('source', {}).get('@timestamp'),
                    'rule_id': a.get('source', {}).get('rule', {}).get('id'),
                    'agent_ip': a.get('source', {}).get('agent', {}).get('ip'),
                    'level': a.get('source', {}).get('rule', {}).get('level')
                }
                for a in alerts_data['results'][:5]
            ]

        # Check recent sent alerts
        from models import SentAlert
        recent_sent = SentAlert.query.filter(
            SentAlert.alert_config_id == alert_id,
            SentAlert.timestamp >= datetime.utcnow() - timedelta(hours=24)
        ).order_by(SentAlert.timestamp.desc()).limit(10).all()

        debug_info['sent_alerts'] = [
            {
                'identifier': sent.alert_identifier[:20] + '...',
                'timestamp': sent.timestamp.isoformat()
            }
            for sent in recent_sent
        ]

        return jsonify(debug_info)

    except Exception as e:
        logger.error(f"Error debugging alert config: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alerts/manual_check', methods=['POST'])
@login_required
def manual_alert_check():
    """Manually trigger alert checking for all enabled configurations"""
    try:
        if not current_user.is_admin():
            return jsonify({"error": "Admin access required"}), 403

        import scheduler

        # Run alert checking manually
        scheduler.check_alerts()

        return jsonify({
            "message": "Manual alert check completed successfully",
            "timestamp": datetime.utcnow().isoformat()
        })

    except Exception as e:
        logger.error(f"Error in manual alert check: {str(e)}")
        return jsonify({"error": str(e)}), 500

@alerts_bp.route('/api/alerts/export', methods=['GET'])
@login_required
def export_alerts():
    """Export alerts in CSV, XLSX, or PDF format"""
    try:
        import csv
        import io
        from flask import make_response
        from opensearch_api import OpenSearchAPI

        opensearch = OpenSearchAPI()

        # Get query parameters (same as get_alerts)
        severity_levels = request.args.getlist('severity_levels[]')
        if not severity_levels:
            severity_levels = ['critical', 'high']

        # Time range
        time_range = request.args.get('time_range', '24h')
        end_time = datetime.utcnow().isoformat()

        if time_range == '1h':
            start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()
        elif time_range == '6h':
            start_time = (datetime.utcnow() - timedelta(hours=6)).isoformat()
        elif time_range == '24h':
            start_time = (datetime.utcnow() - timedelta(days=1)).isoformat()
        elif time_range == '7d':
            start_time = (datetime.utcnow() - timedelta(days=7)).isoformat()
        elif time_range == '30d':
            start_time = (datetime.utcnow() - timedelta(days=30)).isoformat()
        elif time_range == '60d':
            start_time = (datetime.utcnow() - timedelta(days=60)).isoformat()
        elif time_range == '90d':
            start_time = (datetime.utcnow() - timedelta(days=90)).isoformat()
        else:
            start_time = request.args.get('start_time')
            end_time = request.args.get('end_time', end_time)

        # Get export format
        export_format = request.args.get('format', 'csv').lower()

        # Additional filters
        additional_filters = {}
        search_query = request.args.get('search_query')
        if search_query:
            additional_filters['query'] = search_query # Use a generic query field

        rule_id = request.args.get('rule_id')
        if rule_id:
            additional_filters['rule.id'] = rule_id

         # FIM alerts filter
        fim_alerts = request.args.get('fim_alerts')
        if fim_alerts == 'true':
            additional_filters['rule.id'] = ['553', '554']

        # Get all alerts for export (no pagination limit)
        results = opensearch.search_alerts(
            severity_levels=severity_levels,
            start_time=start_time,
            end_time=end_time,
            limit=10000,  # Large limit for export
            offset=0,
            sort_field='@timestamp',
            sort_order='desc',
            additional_filters=additional_filters
        )

        if 'results' not in results:
            return jsonify({'error': 'No alerts found'}), 404

        alerts_data = results['results']

        if export_format == 'csv':
            return export_alerts_csv(alerts_data)
        elif export_format == 'xlsx':
            return export_alerts_xlsx(alerts_data)
        elif export_format == 'pdf':
            try:
                return export_alerts_pdf(alerts_data)
            except ImportError:
                return jsonify({'error': 'PDF export requires reportlab package. Please install it.'}), 500
            except Exception as e:
                 logger.error(f"Error generating PDF report: {str(e)}")
                 flash("Error generating report: PDF generation is not available. System dependencies are missing. Please use HTML format instead.", 'error')
                 return redirect(url_for('alerts.index'))
        else:
            return jsonify({'error': 'Unsupported format'}), 400

    except Exception as e:
        logger.error(f"Error exporting alerts: {str(e)}")
        return jsonify({'error': str(e)}), 500

def export_alerts_csv(alerts_data):
    """Export alerts as CSV"""
    import csv
    import io
    from flask import make_response

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Timestamp',
        'Agent Name',
        'Agent ID', 
        'Agent IP',
        'Rule ID',
        'Rule Description',
        'Severity Level',
        'Location'
    ])

    # Write data
    for alert in alerts_data:
        source = alert.get('source', {})
        timestamp = source.get('@timestamp', 'N/A')
        agent = source.get('agent', {})
        rule = source.get('rule', {})

        # Format timestamp
        try:
            if timestamp != 'N/A':
                formatted_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
            else:
                formatted_time = 'N/A'
        except:
            formatted_time = timestamp

        writer.writerow([
            formatted_time,
            agent.get('name', 'N/A'),
            agent.get('id', 'N/A'),
            agent.get('ip', 'N/A'),
            rule.get('id', 'N/A'),
            rule.get('description', 'N/A'),
            rule.get('level', 'N/A'),
            agent.get('labels', {}).get('location', {}).get('set', 'N/A')
        ])

    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return response

def export_alerts_xlsx(alerts_data):
    """Export alerts as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io
        from flask import make_response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security Alerts"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write headers
        headers = [
            'Timestamp', 'Agent Name', 'Agent ID', 'Agent IP',
            'Rule ID', 'Rule Description', 'Severity Level', 'Location'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_idx, alert in enumerate(alerts_data, 2):
            source = alert.get('source', {})
            timestamp = source.get('@timestamp', 'N/A')
            agent = source.get('agent', {})
            rule = source.get('rule', {})

            # Format timestamp
            try:
                if timestamp != 'N/A':
                    formatted_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    formatted_time = 'N/A'
            except:
                formatted_time = timestamp

            row_data = [
                formatted_time,
                agent.get('name', 'N/A'),
                agent.get('id', 'N/A'),
                agent.get('ip', 'N/A'),
                rule.get('id', 'N/A'),
                rule.get('description', 'N/A'),
                rule.get('level', 'N/A'),
                agent.get('labels', {}).get('location', {}).get('set', 'N/A')
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response

    except ImportError:
        return jsonify({'error': 'Excel export requires openpyxl package. Please install it.'}), 500

def export_alerts_pdf(alerts_data):
    """Export alerts as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import io
        from flask import make_response

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        # Content list
        content = []

        # Title
        title = Paragraph("Security Alerts Export", title_style)
        content.append(title)
        content.append(Spacer(1, 12))

        # Export info
        export_info = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        content.append(export_info)
        content.append(Paragraph(f"Total Alerts: {len(alerts_data)}", styles['Normal']))
        content.append(Spacer(1, 20))

        # Prepare table data
        table_data = [['Timestamp', 'Agent', 'Rule ID', 'Description', 'Level']]

        for alert in alerts_data[:100]:  # Limit to first 100 for PDF
            source = alert.get('source', {})
            timestamp = source.get('@timestamp', 'N/A')
            agent = source.get('agent', {})
            rule = source.get('rule', {})

            # Format timestamp
            try:
                if timestamp != 'N/A':
                    formatted_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%m/%d %H:%M')
                else:
                    formatted_time = 'N/A'
            except:
                formatted_time = timestamp[:10] if len(timestamp) > 10 else timestamp

            # Truncate long descriptions
            description = rule.get('description', 'N/A')
            if len(description) > 40:
                description = description[:37] + '...'

            table_data.append([
                formatted_time,
                agent.get('name', 'N/A')[:15],
                str(rule.get('id', 'N/A')),
                description,
                str(rule.get('level', 'N/A'))
            ])

        # Create table
        table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 0.8*inch, 2.5*inch, 0.6*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        content.append(table)

        if len(alerts_data) > 100:
            content.append(Spacer(1, 12))
            content.append(Paragraph(f"Note: Only first 100 alerts shown. Total: {len(alerts_data)}", styles['Italic']))

        # Build PDF
        doc.build(content)
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return response

    except ImportError:
        return jsonify({'error': 'PDF export requires reportlab package. Please install it.'}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>/test', methods=['POST'])
@login_required
def test_alert(alert_id):
    """Test an alert configuration by sending a test email"""
    try:
        from models import AlertConfig
        from opensearch_api import OpenSearchAPI
        from email_alerts import EmailAlerts
        
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        # Create email sender
        email_alerts = EmailAlerts()

        # Get some real alert data for the test
        opensearch = OpenSearchAPI()
        severity_levels = alert.get_alert_levels()

        # Get alerts from the last hour for testing
        end_time = datetime.utcnow().isoformat()
        start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()

        alerts_data = opensearch.search_alerts(
            severity_levels=severity_levels,
            start_time=start_time,
            end_time=end_time,
            limit=5  # Just get a few for the test
        )

        # Send the real alert email
        if alerts_data and 'results' in alerts_data and len(alerts_data['results']) > 0:
            success = email_alerts.send_severity_alert(alert, alerts_data)
            message = f"Test alert sent successfully using real alert data to {alert.email_recipient}"
        else:
            # If no real alerts found, just send a basic test message
            subject = f"Test Alert from AZ Sentinel X - {alert.name}"

            # Get include fields if available
            include_fields = []
            if hasattr(alert, 'get_include_fields') and callable(getattr(alert, 'get_include_fields')):
                include_fields = alert.get_include_fields()

            fields_html = ""
            if include_fields:
                fields_html = f"""
                <p>Included Fields:</p>
                <ul>
                    {''.join([f'<li>{field}</li>' for field in include_fields])}
                </ul>
                """

            message = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; }}
                    .alert-box {{ padding: 15px; background-color: #f8f9fa; border-left: 5px solid #28a745; margin-bottom: 20px; }}
                    ul {{ margin-top: 5px; }}
                    li {{ margin-bottom: 3px; }}
                </style>
            </head>
            <body>
                <h1>Test Alert</h1>
                <div class="alert-box">
                    <p>This is a test alert from AZ Sentinel X.</p>
                    <p>Alert Name: {alert.name}</p>
                    <p>Alert Levels: {', '.join(alert.get_alert_levels())}</p>
                    <p>Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    {fields_html}
                </div>
                <p>This is a test email to verify that your alert configuration is working correctly.</p>
                <p>No actual alerts matching your criteria were found in the last hour, so no real data is displayed.</p>
                <p>When actual alerts are found, they will be displayed in a table format with the fields you have selected.</p>
                <p>The system will automatically track and prevent duplicate alerts with the same @timestamp, agent.ip, agent.name, rule.description, and rule.id values.</p>
            </body>
            </html>
            """

            success = email_alerts.send_alert_email(alert.email_recipient, subject, message)
            message = f"Test alert sent successfully to {alert.email_recipient} (no real alerts found)"

        # Check if the email was sent successfully
        if success:
            return jsonify({
                'message': f'Test alert sent successfully to {alert.email_recipient}'
            })
        else:
            return jsonify({
                'error': 'Failed to send test alert email'
            }), 500
    except Exception as e:
        logger.error(f"Error sending test alert: {str(e)}")
        return jsonify({'error': str(e)}), 500